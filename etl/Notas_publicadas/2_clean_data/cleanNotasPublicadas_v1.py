# -*- coding: utf-8 -*-

import os
import csv
from datetime import datetime
import warnings
from openpyxl import load_workbook

# Suprimir avisos do openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Caminho da pasta onde os arquivos .xlsx estao localizados
pasta_downloads = r"/home/geraldo.junior/Database_mkt/etl/Downloads"

# Caminho do arquivo CSV de saida
output_file_path = r"/home/geraldo.junior/Database_mkt/etl/Downloads/notas_limpas.csv"

# Colunas que devem ser mantidas
colunas_mantidas = [
    "id_nota", 
    "sacado", 
    "vl_documento", 
    "dt_inclusao", 
    "cedente", 
    "cnpj_cedente", 
    "dt_vcto", 
    "dt_emissao"
]

# Funcao para formatar datas no formato YYYY-MM-DD
def formatar_data(data):
    if isinstance(data, datetime):
        return data.strftime("%Y-%m-%d")
    if isinstance(data, str):
        for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(data, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None

# Funcao para converter valores numericos
def converter_valores(item):
    try:
        item["cnpj_cedente"] = int(float(item["cnpj_cedente"])) if item["cnpj_cedente"] else None
    except:
        item["cnpj_cedente"] = None

    try:
        item["vl_documento"] = float(item["vl_documento"]) if item["vl_documento"] else None
    except:
        item["vl_documento"] = None

    try:
        item["id_nota"] = int(float(item["id_nota"])) if item["id_nota"] else None
    except:
        item["id_nota"] = None

    return item

# Funcao para limpar os dados
def limpar_csv(data, colunas_mantidas):
    dados_limpos = []
    for item in data:
        item_limpo = {col: item.get(col) for col in colunas_mantidas}
        for col_data in ["dt_vcto", "dt_inclusao", "dt_emissao"]:
            item_limpo[col_data] = formatar_data(item_limpo.get(col_data))
        item_limpo = converter_valores(item_limpo)
        dados_limpos.append(item_limpo)
    return dados_limpos

# Listar arquivos .xlsx
arquivos_xlsx = [arq for arq in os.listdir(pasta_downloads) if arq.endswith('.xlsx')]
if not arquivos_xlsx:
    print("Nenhum arquivo .xlsx encontrado.")
    exit()

# Carrega o primeiro arquivo
input_file_path = os.path.join(pasta_downloads, arquivos_xlsx[0])

# Leitura do Excel com openpyxl
try:
    wb = load_workbook(input_file_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    dados = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        dados.append(item)
except Exception as e:
    print(f"Erro ao ler o arquivo Excel: {e}")
    exit()

# Limpar os dados
dados_limpos = limpar_csv(dados, colunas_mantidas)

# Agrupar GM mantendo dt_inclusao mais recente
dados_gm = [d for d in dados_limpos if d.get("sacado") == "GM - GENERAL MOTORS DO BRASIL LTDA"]
dados_gm_agrupados = {}

for item in dados_gm:
    chave = tuple(item[c] for c in ["vl_documento", "cnpj_cedente", "dt_vcto", "dt_emissao"])
    dt_inclusao = item.get("dt_inclusao") or ''
    if chave not in dados_gm_agrupados or dt_inclusao > (dados_gm_agrupados[chave].get("dt_inclusao") or ''):
        dados_gm_agrupados[chave] = item

# Outros dados (nao GM)
dados_outros = [d for d in dados_limpos if d.get("sacado") != "GM - GENERAL MOTORS DO BRASIL LTDA"]

# Combina resultados
dados_finais = dados_outros + list(dados_gm_agrupados.values())

# Salva o CSV
try:
    with open(output_file_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=colunas_mantidas, delimiter=';')
        writer.writeheader()
        writer.writerows(dados_finais)
    print(f"Arquivo CSV limpo salvo com sucesso em: {output_file_path}")
except Exception as e:
    print(f"Erro ao salvar o arquivo CSV: {e}")
